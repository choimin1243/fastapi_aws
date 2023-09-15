from fastapi import  UploadFile, File
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook  # 추가된 부분
import re
import pandas as pd
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi import FastAPI, Form, Request
from fastapi.responses import StreamingResponse
import openpyxl
import json
from fastapi import APIRouter
from fastapi.templating import Jinja2Templates
from fastapi import FastAPI, Form, Request, Depends
from fastapi import APIRouter,Request
from package import get_all_todos_from_db,hello
from database import SessionLocal  # SessionLocal을 불러옴
db = SessionLocal()



appends = APIRouter()
templates = Jinja2Templates(directory="templates")
templates.env.globals.update(enumerate=enumerate)

def flatten(lst):
    result = []
    for item in lst:
        if isinstance(item, list):
            result.extend(flatten(item))
        else:
            result.append(item)
    return result


def filterate(listly):
    database_list=hello(db)
    listly=flatten(listly)
    list_result=' '.join(map(str, listly))
    x=""
    for i in range(len(database_list)):
        finder=database_list[i]
        if finder in list_result:
            x= finder
            break

    return x

# HTML 템플릿을 렌더링하는 엔드포인트
@appends.get("/")
async def render_upload_form(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})




# @app.post("/send-list/")
# async def send_list(request: Request, selected_columns: str = Form(...), content_items: str = Form(...), encoded_data: str = Form(...)):
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet["A1"] = "Hello"
#     # ... 데이터 추가 작업 ...
#
#     # 엑셀 워크북을 바이트로 저장
#     output_excel = io.BytesIO()
#     workbook.save(output_excel)
#     output_excel.seek(0)  # 파일 위치를 처음으로 되돌림
#
#     # 메모리에 생성된 엑셀 파일을 스트리밍하여 사용자에게 전송
#     return FileResponse("교원.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="output.xlsx")


def wat_resize(wat):
    패턴 = r"(\d+)/(\d+)"
    매치 = re.search(패턴, wat)
    pp = r'\b(\d+)w\b'
    pattern = r'\b(\d+)\s*(?:밀리와트|mW|mw)\b'
    matcheses = re.findall(pattern, wat, re.IGNORECASE)

    패턴즈 = r"kw"
    m=r'mw'
    매치들 = re.findall(패턴즈, wat, re.IGNORECASE)
    if 매치들:
        if 매치:
            분자 = int(매치.group(1))
            분모 = int(매치.group(2))
            소수 = 분자 / 분모
            결과_문자열 = float(소수)*1000

            return 결과_문자열

        else:
            패턴 = r"\d+"
            추출된_숫자들 = re.findall(패턴, wat, re.IGNORECASE)
            return float(추출된_숫자들[0])*1000


    와트매치=re.findall(pp, wat, re.IGNORECASE)
    if 와트매치:
        print(wat,"@@)")
        if 매치:
            분자 = int(매치.group(1))
            분모 = int(매치.group(2))
            소수 = 분자 / 분모
            결과_문자열 = float(소수)

            print(결과_문자열)
            return 결과_문자열

        else:
            패턴 = r"\d+"
            추출된_숫자들 = re.findall(패턴, wat, re.IGNORECASE)
            return 추출된_숫자들[0]
    m=r'mw'
    MW=re.findall(m, wat, re.IGNORECASE)

    if MW:
        if 매치:
            분자 = int(매치.group(1))
            분모 = int(매치.group(2))
            소수 = 분자 / 분모
            print("~~~",소수)
            return float(소수)*0.001

        else:
            패턴 = r"\d+"
            추출된_숫자들 = re.findall(패턴, wat, re.IGNORECASE)
            return float(추출된_숫자들[0])*0.001




def om(rest):
    pattern = r"(?<!\w)(\d+)\s*Ω"
    patternk = r'(\d+\.\d+|\d+)(?=\s*[KΩ|㏀])'
    patternm = r'\d+\s*m[Ω|Ω]'
    patternss = r'\d+\s*[^mM]㏁|\d+\s*MΩ'
    if re.search(pattern,rest):
        print(re.findall(pattern,rest))
        return float(re.findall(pattern,rest)[0])*0.001
    if re.search(patternk,rest):
        return float(re.findall(patternk, rest)[0])
    if re.search(patternm,rest):
        patternr = r"\d+"
        print(re.findall(patternr,rest)[0])
        return float(re.findall(patternr,rest)[0])*0.000001
    if re.search(patternss, rest):
        patternr = r"\d+"
        print(rest)
        return float(re.findall(patternr,rest)[0])*1000




def package(sheet):
    print(sheet)











@appends.post("/send-list/")
async def send_list(request: Request, selected_columns: str = Form(...),content_items:str = Form(...), encoded_data:str=Form(...)):
    selected_columns = json.loads(selected_columns)
    content_items=json.loads(content_items)
    encoded_data=json.loads(encoded_data)


    number_to=len(content_items)
    output_excel = BytesIO()
    work = Workbook()


    for charact in content_items:
        data_list = encoded_data# JSON 문자열을 리스트로 변환
        work.create_sheet(title=charact)




        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 데이터를 시트에 추가
        for row_data in data_list:
            sheet.append(row_data)

        last_column = sheet.max_column
        last_row = sheet.max_row

        data = []
        location_column_index = selected_columns
        part_number = []

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
            # 한 행의 데이터를 저장할 리스트를 생성합니다.
            row_data = []
            for idx, cell in enumerate(row, 1):
                row_data.append([cell.value])
            data.append(row_data)






        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
            # 한 행의 데이터를 저장할 리스트를 생성합니다.
            for idx, cell in enumerate(row, 1):
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.lower()
                    if cell.value == "package":
                        part_number = idx


        if part_number == []:
            part_number=500



        def remove_duplicates(input_list):
            return list(set(input_list))

        patternwat = r"(\d+(?:/\d+)?(?:\.\d+)?)\s*(w|kw|mw)"
        patternnp = r"(?<!\w)(?<!\d)\d+(?:\.\d+)?(?:\s*(?:pF|nF|uF|µF|UF|p|n|u|µ))(?!\w)"
        patternv = r"(\d+(?:\.\d+)?)\s*(?:[kK]?[mM]?[vV])"

        tolerance_values = ["J", "F", "A", "B", "G", "M", "Z"]

        # 수정된 정규식
        pattern_tor = r"(?<![A-Za-z0-9.,-])(?:{})(?![A-Za-z0-9.,])".format("|".join(tolerance_values))

        patternAed = r"([-+]?\d+(?:\.\d+)?)\s*([mµ]?[AaKk])"
        patterntemp = r'\d+(?:\.\d+)?\s*℃'

        location_column_index = remove_duplicates(location_column_index)
        pattern_kv = r"kv"
        pattern_v = r"\d+(?:\.\d+)?(?=\s*(?i)v)"

        result_data = []

        character = charact
        pattern = r"(?<!\S)" + character + "(\d+)"

        list_row = []

        for i in range(len(data)):
            try:
                datas = data[i][int(location_column_index[0])]
                if datas[0] != None:
                    parsed_data = [row.replace(" ", "").split(",") for row in datas[0].split("\n")]
                    flattened_data = [item for sublist in parsed_data for item in sublist]
                    if re.findall(pattern, flattened_data[0], re.IGNORECASE):
                        list_row.append(i)
                        for s in range(len(flattened_data)):
                            if flattened_data[s] != '':
                                result_data.append([i, flattened_data[s].strip()])


            except:
                pass

        voltage_number = 1
        wat_number = 1
        resistance_number = 1
        tolerance_number = 1
        nlp = 1
        part_num = 1
        list_table_number = ["No", "REF NO"]

        pattern_kv = r"kv"
        pattern_v = r"(\d+(?:\.\d+)?)\s*(?:[vV])"
        patternv = r"(\d+(?:\.\d+)?)\s*(?:[kK]?[mM]?[vV])"

        for i in range(len(list_row)):
            data_item = data[list_row[i]]

            for s in range(len(data_item)):
                try:
                    something = re.search(patternv, data_item[s][0])
                    voltage_value = something.group(0)
                    for k in range(len(result_data)):
                        if result_data[k][0] == list_row[i]:
                            voltage_value = something.group(0)
                            matches = re.search(pattern_kv, voltage_value, re.IGNORECASE)
                            match = re.search(pattern_v, voltage_value, re.IGNORECASE)

                            if match:
                                voltage_number = 2
                                matches_data = re.findall(pattern_v, voltage_value)
                                result_data[k].append(float(matches_data[0]))
                            if matches:
                                voltage_number = 2
                                matches_num = re.findall(patternv, voltage_value, re.IGNORECASE)
                                result_data[k].append(float(matches_num[0])*1000)

                    break

                except:
                    pass


        if voltage_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("VOLTAGE")


        for i in range(len(list_row)):
            data_item = data[list_row[i]]

            for s in range(len(data_item)):
                try:
                    matches = re.findall(patternwat, data_item[s][0], re.IGNORECASE)
                    if matches:
                        wat_number = 2

                        combined_values = [f"{match[0]}{match[1].lower()}" for match in matches]
                        combined_result = " ".join(combined_values)
                        wat = combined_result
                        for k in range(len(result_data)):
                            patternkw = r"(\w+)\d*KW"
                            patternrw = r"(\w+)\d*W"
                            matchkw = re.search(patternkw, wat, re.IGNORECASE)
                            matchw=re.search(patternrw, wat, re.IGNORECASE)
                            numfrac = r"\d+/\d+"

                            if result_data[k][0] == list_row[i]:
                                a=wat_resize(wat)
                                result_data[k].append(a)

                        break
                    else:
                        wat = ""

                except:
                    wat = ""

        if wat_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("RATED_POWER[W]")

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(pattern_tor, data_item[s][0])
                    if match:
                        tolerance_number = 2
                        tolerance_value = match.group(0)
                        print(tolerance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(tolerance_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if tolerance_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("TOLERANCE")

        patternom = r"(?:,\s*)?(\d+(?:\.\d+)?)(?:\s*(?:㏀|Ω|k㏀|kΩ|mΩ|㏁|MΩ))\s*\*?\d?"

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    matchnorm = re.search(patternom, data_item[s][0])

                    if matchnorm:
                        resistance_number = 2
                        resistance_value = matchnorm.group(0)
                        patternmega = r'㏁'
                        patternmili= r'mΩ'
                        print("~",resistance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                pattern =r"(?<!\w)(\d+)\s*Ω"
                                patterned = r'[KΩ|㏀]'
                                result_data[k].append(om(resistance_value))

                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if resistance_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("RESISTANCE")


        temp_number = 1

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(patterntemp, data_item[s][0], re.IGNORECASE)
                    if match:
                        temp_number = 2
                        tmp_value = match.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                pattern = r'\d+'
                                number = re.findall(pattern, tmp_value)

                                result_data[k].append(float(number[0]))
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if temp_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("TEMPERATURE")

        #
        nlp_number = 1
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(patternnp, data_item[s][0], re.IGNORECASE)
                    if match:
                        nlp_number = 2
                        nlp_value = match.group(0)
                        print(tolerance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(nlp_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if nlp_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("CAPACITANCE")
        #
        #
        #
        #
        if part_number!=500:
            for i in range(len(list_row)):
                data_item = data[list_row[i]]
                for k in range(len(result_data)):
                    if result_data[k][0] == list_row[i]:
                        result_data[k].append(data_item[part_number - 1][0])
                        
                        
        if part_number==500:
            for i in range(len(list_row)):
                data_item = data[list_row[i]]
                for k in range(len(result_data)):
                    if result_data[k][0] == list_row[i]:
                        strings=filterate(data_item)
                        print(k,strings)
                        result_data[k].append(strings)


        print("test",result_data)

        print("end@@@")
        list_table_number.append("PACKAGE")

        pattern_caps = r"(X7R|X5R|COG|NPO|X5S|X6S|C0G)"

        Grade = 1
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(pattern_caps, data_item[s][0], re.IGNORECASE)
                    if match:
                        Grade = 2
                        Grade_value = match.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(Grade_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if Grade == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("GRADE")

        for row in result_data:
            if row[1].isdigit():
                row[1] = character + row[1]


        for row in result_data:
            num = len(character)
            row[0] = int(row[1][num:])

        result_data.insert(0, list_table_number)


        df = pd.DataFrame(result_data[1:], columns=result_data[0])

        # 첫 번째 열을 기준으로 정렬

        A_table = ["No", "REF NO", "PACKAGE", "RATED_POWER[W]", "TOLERANCE", "RESISTANCE"]
        B_table = ["No", "REF NO", "PACKAGE", "CAPACITANCE", "VOLTAGE", "GRADE", "TOLERANCE", "TEMPERATURE"]

        sorted_df = df.sort_values(by='No')


        print("@@@",sorted_df)
        if character == "R":
            column_order = A_table
        else:
            column_order = B_table

        for column in column_order:
            if column not in sorted_df:
                sorted_df[column] = float("nan")  # 모든 값은 NaN으로 설정합니다.





        if character=="R":
            sorted_df.loc[sorted_df['PACKAGE'] == '0402', 'RESISTANCE'] = 0.031
            sorted_df.loc[sorted_df['PACKAGE'] ==int('0402'), 'RESISTANCE'] = 0.031

            sorted_df.loc[sorted_df['PACKAGE'] == '0603', 'RESISTANCE'] = 0.05
            sorted_df.loc[sorted_df['PACKAGE'] == int('0603'), 'RESISTANCE'] = 0.05

            sorted_df.loc[sorted_df['PACKAGE'] == '1005', 'RESISTANCE'] = 0.0625
            sorted_df.loc[sorted_df['PACKAGE'] == int('1005'), 'RESISTANCE'] = 0.0625

            sorted_df.loc[sorted_df['PACKAGE'] == '1608', 'RESISTANCE'] = 0.1
            sorted_df.loc[sorted_df['PACKAGE'] == int('1608'), 'RESISTANCE'] = 0.1

            sorted_df.loc[sorted_df['PACKAGE'] == '2012', 'RESISTANCE'] = 0.125
            sorted_df.loc[sorted_df['PACKAGE'] == int('2012'), 'RESISTANCE'] = 0.125

            sorted_df.loc[sorted_df['PACKAGE'] == '3216', 'RESISTANCE'] = 0.25
            sorted_df.loc[sorted_df['PACKAGE'] == int('3216'), 'RESISTANCE'] = 0.25

            sorted_df.loc[sorted_df['PACKAGE'] == '3225', 'RESISTANCE'] = 0.5
            sorted_df.loc[sorted_df['PACKAGE'] == int('3225'), 'RESISTANCE'] = 0.5

            sorted_df.loc[sorted_df['PACKAGE'] == '6432', 'RESISTANCE'] = 1
            sorted_df.loc[sorted_df['PACKAGE'] == int('6432'), 'RESISTANCE'] = 1

        else:
            pass

        print(sorted_df)
        sorted_df_by_column_order = sorted_df[column_order]
        charact_sheet = work[character]
        for row in dataframe_to_rows(sorted_df_by_column_order, index=False, header=True):
            charact_sheet.append(row)










    work.save(output_excel)

    # BytesIO의 파일 포인터를 처음으로 이동시킴
    output_excel.seek(0)

    return StreamingResponse(output_excel,
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=result.xlsx"})


@appends.post("/upload/")
async def upload_excel_file(request: Request,file: UploadFile = File(...)):
    file_path =file.filename
    print(file_path)
    with open(file_path, "wb") as temp_file:
        temp_file.write(file.file.read())


    file_data = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        file_data.append(row)


    encoded_file_data = json.dumps(file_data)  # Convert the list to JSON string

    # 결과를 템플릿에 전달하여 렌더링
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "file_data": file_data, "encoded_data":encoded_file_data, "file_path":file_path}
    )


